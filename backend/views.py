import json
import pickle
import re

import numpy as np
from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse,JsonResponse
import django.http as http
import os
from .config import *
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from model.pred import getResultOfDataset_wensi,getResultOfDataset_GM,getResultWithParams_wensi,getResultWithParams_GM,getResultWithParams_prophet,getResultOfDataset_prophet
import pymongo
from .util import *
from data.preprocess import Method1
import pandas
import openpyxl




def index(request):
    return HttpResponse("Hello, world. You're at the polls index.")


@csrf_exempt
def upload(request):
    if request.method == 'POST':
        if request.FILES:
            myFile = None
            for i in request.FILES:
                myFile = request.FILES[i]
            if myFile:
                dir = BASE_DIR
                destination = open(os.path.join(dir, myFile.name),
                                   'wb+')
                for chunk in myFile.chunks():
                    destination.write(chunk)
                destination.close()
            return HttpResponse('ok')
        else:
            return HttpResponse("none")
    else:
        return HttpResponse("upload error")


@csrf_exempt
def getAllMetaModels(request):
    models = ['prophet',"翁氏模型","灰度预测"]
    resp = [{"value":item,"label":item} for item in models]
    return JsonResponse(resp,safe=False)

@csrf_exempt
def getAllDatasets(request):
    all_datasets = []
    files = os.listdir(BASE_DIR)
    for file in files:
        if file == "all_dataset_tag.pkl":
            continue
        sheets = pd.read_excel(os.path.join(BASE_DIR,file), sheet_name=None)
        for sheet in sheets:
            all_datasets.append(file.split(".")[0]+"_"+sheet)

    resp = [{"value": item, "label": item} for item in all_datasets]
    return JsonResponse(resp,safe=False)

@csrf_exempt
def getAllPreprocessMethods(request):
    resp = [
        {"value":"平滑处理","label":"平滑处理"},
    ]
    return JsonResponse(resp,safe=False)

@csrf_exempt
def getResultOfPreprocess(request):
    dataset = request.GET.get('dataset', '')
    method = request.GET.get('method', '')
    window_size = [int(item) for item in list(request.GET.get('window_size','').split(","))]
    print("dataset:",dataset,"method:",method,"window_size",window_size)

    fileName, sheetName = getFileName(dataset)
    data = pd.read_excel(os.path.join(BASE_DIR, fileName), header=0, skiprows=0,
                         sheet_name=sheetName).to_numpy().transpose().tolist()

    obj = {
        "dataset_xAxis": data[0],
        "dataset_yAxis": data[1],
    }
    for size in window_size:
        _data = Method1(dataset,size)
        obj[method+str(size)]=_data[1]
    return JsonResponse(obj,safe=False)

@csrf_exempt
def saveDataset(request):
    input = json.loads(request.body.decode('utf-8'))
    name = input.get('name')
    data = input.get('data')
    base_data = input.get("base_data")
    window_size = input.get("window_size")
    print("name:",name)
    print("data:",data)
    print("base_data:",base_data)
    print("window_size",window_size)

    if len(data.keys())==0 or len(base_data.keys())==0:
        print("数据为空，保存失败")
        return JsonResponse(name, safe=False)

    # 首先判断是否存在当前sheet
    xl = pd.ExcelFile(os.path.join(BASE_DIR,name.split("_")[0]+".xlsx"))
    print(xl.sheet_names)
    if "_".join(name.split("_")[1:]) in xl.sheet_names:
        print("预处理数据已存在")
        return JsonResponse(name, safe=False)
    # 保存文件
    # 创建一个Workbook对象，相当于创建了一个Excel文件
    workbook=openpyxl.load_workbook(os.path.join(BASE_DIR,name.split("_")[0]+".xlsx"))
    worksheet2 = workbook.create_sheet()
    worksheet2.title = "_".join(name.split("_")[1:])
    col1 = ["年份"] + base_data["xAxis"]
    col2 = ["实际值"] + data["".join(name.split("_")[-2:])]

    for i in range(0, len(col1)):
        worksheet2.cell(i + 1, 1, col1[i])

    # 填入第二列
    for i in range(0, len(col2)):
        worksheet2.cell(i + 1, 2, col2[i])

    workbook.save(os.path.join(BASE_DIR,name.split("_")[0]+".xlsx"))
    #workbook.save(os.path.join(BASE_DIR, "test" + ".xlsx"))
    print(worksheet2.values)

    return JsonResponse(name,safe=False)

@csrf_exempt
def getResultOfModel(request):
    if request.method=='POST':
        data = json.loads(request.body.decode('utf-8'))
        models = data.get('models')
        dataset_name = data.get('dataset')
        print("models:",models)
        print("dataset:",dataset_name)

        fileName = None
        for file in os.listdir(BASE_DIR):
            if file.split(".")[0] == dataset_name.split("_")[0]:
                fileName = file
                break
        sheetName = dataset_name.split("_")[1]
        data = pd.read_excel(os.path.join(BASE_DIR, fileName), header=0, skiprows=0,sheet_name=sheetName).to_numpy().transpose().tolist()

        obj = {
            "dataset_xAxis":data[0],
            "dataset_yAxis":data[1]
        }
        for model in models:
            if model=="prophet":
                obj["prophet"] = getResultOfDataset_prophet(dataset_name)
            elif model=="翁氏模型":
                obj["翁氏模型"] = getResultOfDataset_wensi(dataset_name)
            elif model=="灰度预测":
                obj["灰度预测"] = getResultOfDataset_GM(dataset_name)


        print(obj)
        return JsonResponse(
            obj,safe=False
        )


@csrf_exempt
def getResultWithParams(request):
    if request.method=='POST':
        data = json.loads(request.body.decode('utf-8'))
        model = data.get('model')
        dataset = data.get('dataset')
        params = data.get("params")

        print("model:",model)
        print("dataset:",dataset)
        print("params:",params)
        fileName,sheetName = getFileName(dataset)
        data = pd.read_excel(os.path.join(BASE_DIR, fileName), header=0, skiprows=0,sheet_name=sheetName).to_numpy().transpose().tolist()

        if params["years"]>0:
            for i in range(params["years"]):
                data[0].append(data[0][-1]+1)

        obj = {
            "dataset_xAxis":data[0],
            "dataset_yAxis":data[1]
        }

        def get_loss(a,b):
            sum = 0
            for i in range(len(a)):
                sum +=(a[i]-b[i])*(a[i]-b[i])
            return sum/len(a)

        if model=="prophet":
            obj["prophet"],obj["k"] = getResultWithParams_prophet(dataset,params)
            print(obj["k"],type(obj["k"]))
            obj["k"] = float(obj["k"])
            obj["loss"] = get_loss(obj["dataset_yAxis"],obj["prophet"])

        elif model=="翁氏模型":
            obj["翁氏模型"],obj["a"],obj["b"],obj["c"] = getResultWithParams_wensi(dataset,params)
            obj["loss"] = get_loss(obj["dataset_yAxis"], obj["翁氏模型"])
        elif model=="灰度预测":
            obj["灰度预测"],msg = getResultWithParams_GM(dataset,params)
            obj["loss"] = get_loss(obj["dataset_yAxis"], obj["灰度预测"])
            if msg != None:
                obj["msg"] = msg

        print("obj",obj)
        return JsonResponse(
            obj,safe=False
        )


@csrf_exempt
def saveModel(request):
    if request.method=='POST':
        data = json.loads(request.body.decode('utf-8'))

        model = data.get('model')
        dataset = data.get('dataset')
        params = data.get("params")


        print("model:",model)
        print("dataset:",dataset)
        print("params:",params)

        if model=="prophet":
            saveModelToMongo_prophet(params,dataset)
        elif model=="翁氏模型":
            saveModelToMongo_wensi(params,dataset)
        elif model=="灰度预测":
            saveModelToMongo_GM(params,dataset)
        return HttpResponse("ok")


def getDataset(request):
    query = request.GET.get('dataset', '')
    if query.split("_")[0] not in [file.split(".")[0] for file in os.listdir(BASE_DIR)]:
        print(query)
        return  HttpResponse("数据集不存在")

    fileName,sheetName = getFileName(query)

    data = pd.read_excel(os.path.join(BASE_DIR,fileName), header=0, skiprows=0,sheet_name=sheetName).to_numpy().transpose().tolist()
    print(data)
    cumulative = [data[1][0]]
    for i in range(1,len(data[1])):
        cumulative.append(data[1][i]+cumulative[-1])
    return JsonResponse(
        {
            "name":query,
            "xAxis":data[0],
            "yAxis":data[1],
            "cumulative":cumulative,
        }
    )


@csrf_exempt
def saveTag(request):
    dataset = request.GET.get('dataset', '')
    year = request.GET.get('year','')
    print("saveTag",dataset,year)
    # 将数据直接保存为json 文件
    if not os.path.exists(os.path.join(BASE_DIR,'all_dataset_tag.pkl')):
        all_dataset_tag = {}
    else:
        all_dataset_tag = pickle.load(open(os.path.join(BASE_DIR,'all_dataset_tag.pkl'),"rb"))
    if dataset not in all_dataset_tag.keys():
        all_dataset_tag[dataset] = [year]
    else:
        # 判断是否已经存在，如果已经存在，就去掉
        if year in  all_dataset_tag[dataset]:
            all_dataset_tag[dataset].remove(year)
        else:
            all_dataset_tag[dataset].append(year)
            all_dataset_tag[dataset].sort()
    pickle.dump(all_dataset_tag,open(os.path.join(BASE_DIR,"all_dataset_tag.pkl"),"wb"))
    return HttpResponse("ok")

@csrf_exempt
def getTagData(request):
    all_dataset_tag = pickle.load(open(os.path.join(BASE_DIR, 'all_dataset_tag.pkl'), "rb"))
    print("all tags",all_dataset_tag)

    dataset = request.GET.get('dataset', '')
    cur_tag = request.GET.get('curtag', '')
    print("getTagData",dataset)
    if not os.path.exists(os.path.join(BASE_DIR,'all_dataset_tag.pkl')):
        return JsonResponse({
            'data':[],
        })
    else:
        all_dataset_tag = pickle.load(open(os.path.join(BASE_DIR,'all_dataset_tag.pkl'),"rb"))
        if dataset not in all_dataset_tag.keys() or cur_tag not in all_dataset_tag[dataset].keys():
            return JsonResponse({
                'data': [],
            })
        else:
            return JsonResponse({
                'data': all_dataset_tag[dataset],
            })



@csrf_exempt
def getModelList(request):
    model = request.GET.get("model",'')
    myclient = pymongo.MongoClient("mongodb://localhost:27017/")
    mydb = myclient["lab3"]
    if model=="prophet":
        mycol = mydb["prophet"]
    elif model=="翁氏模型":
        mycol = mydb["翁氏模型"]
    elif model=="灰度预测":
        mycol = mydb["灰度预测"]

    res = []
    for x in mycol.find():
        if x["name"]==None:
            continue
        res.append(x["name"])
    
    resp = [{"value":item,"label":item} for item in res]
    return JsonResponse(resp,safe=False)


@csrf_exempt
def loadModel_multi(request):
    data = json.loads(request.body.decode('utf-8'))
    model = data.get("models")
    years = data.get("years")
    print("models:",model)
    print("years:",years)

    obj = {}
    if model["prophet"][0]!="":
        res = loadModel_mutli_sub(model["prophet"][0],"prophet",years)
        obj.update(res)
    if model["灰度预测"][0] != "":
        res = loadModel_mutli_sub(model["灰度预测"][0], "灰度预测", years)
        obj.update(res)
    if model["翁氏模型"][0] != "":
        res = loadModel_mutli_sub(model["翁氏模型"][0], "翁氏模型", years)
        obj.update(res)

    return JsonResponse(obj,safe=False)





@csrf_exempt
def loadModel(request):
    data = json.loads(request.body.decode('utf-8'))
    model = data.get("model")
    name = data.get('name')
    years = data.get("years")
    myclient = pymongo.MongoClient("mongodb://localhost:27017/")
    mydb = myclient["lab3"]
    if model == "prophet":
        mycol = mydb["prophet"]
    elif model == "翁氏模型":
        mycol = mydb["翁氏模型"]
    elif model == "灰度预测":
        mycol = mydb["灰度预测"]

    res = None
    for x in mycol.find():
        if x["name"]==name:
            res = x
    
    
    fileName,sheetName = getFileName(res["dataset"])
    data = pd.read_excel(os.path.join(BASE_DIR, fileName), header=0, skiprows=0,sheet_name=sheetName).to_numpy().transpose().tolist()

    if years>0:
        for i in range(years):
            data[0].append(data[0][-1]+1)

    obj = {
            "dataset_xAxis":data[0],
            "dataset_yAxis":data[1]
        }

    res["years"] = years

    if model=="prophet":
        obj["prophet"],obj["k"] = getResultWithParams_prophet(res["dataset"],res)
        obj["k"] = float(obj["k"])
    elif model=="翁氏模型":
        obj["翁氏模型"],obj["a"],obj["b"],obj["c"] = getResultWithParams_wensi(res["dataset"],res)
    elif model=="灰度预测":
        obj["灰度预测"],msg = getResultWithParams_GM(res["dataset"],res)
        if msg!=None:
            obj["msg"] = msg

    print(obj)


    return JsonResponse(
            obj,safe=False
    )



if __name__=='__main__':
    data = pd.read_excel("../data/datasets/三个样本.xlsx",header=0, skiprows=0)
    print(data)

